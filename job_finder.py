"""Job Finder — Apify-powered scraper for India internships.

Flow:
  1. Load user_profile.yaml + resume PDF.
  2. Call configured Apify actors in parallel
     (LinkedIn Jobs, LinkedIn Posts, Indeed India, Naukri, Internshala, Wellfound, Glassdoor).
  3. Filter results: India-only, no excluded keywords, deduplicate by URL.
  4. Export jobs.xlsx — formatted for upload to Claude AI.
  5. Save <name>_claude_prompt.txt — paste directly into Claude.ai.

Usage:
  python job_finder.py
  python job_finder.py --resume "path/to/resume.pdf"
  python job_finder.py --profile user_profile.yaml --output jobs.xlsx
"""

import argparse
import asyncio
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote_plus

import httpx
import openpyxl
import yaml
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from PyPDF2 import PdfReader

load_dotenv()

APIFY_BASE = "https://api.apify.com/v2"
APIFY_TOKEN = os.getenv("APIFY_TOKEN", "")
OUTPUT_EXCEL = os.getenv("OUTPUT_EXCEL", "jobs.xlsx")
POLL_INTERVAL = 8    # seconds between run-status checks
RUN_TIMEOUT = 420    # max seconds to wait for one actor run
MAX_PARALLEL = 4     # concurrent Apify runs

# ---------------------------------------------------------------------------
# Location filter keywords
# ---------------------------------------------------------------------------
INDIA_LOCS = {
    "india", "delhi", "new delhi", "mumbai", "bangalore", "bengaluru",
    "hyderabad", "chennai", "pune", "kolkata", "noida", "gurugram",
    "gurgaon", "ahmedabad", "jaipur", "bhopal", "indore", "kochi",
    "remote", "work from home", "wfh", "pan india", "anywhere in india",
}
NON_INDIA_LOCS = {
    "usa", "united states", "u.s.", " us ", "uk", "united kingdom",
    "canada", "australia", "germany", "france", "singapore", "dubai",
    "uae", "netherlands", "sweden", "japan", "china", "korea", "brazil",
    "new york", "san francisco", "london", "toronto", "sydney",
}


# ---------------------------------------------------------------------------
# Profile & resume helpers
# ---------------------------------------------------------------------------
def load_profile(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def read_resume(path: str) -> str:
    p = Path(path)
    if not p.exists():
        print(f"  Warning: resume not found at {path}")
        return ""
    reader = PdfReader(str(p))
    return "".join(page.extract_text() or "" for page in reader.pages)


# ---------------------------------------------------------------------------
# Apify API helpers
# ---------------------------------------------------------------------------
async def _start_run(
    client: httpx.AsyncClient, actor_id: str, payload: dict
) -> str | None:
    try:
        api_id = actor_id.replace("/", "~")
        r = await client.post(
            f"{APIFY_BASE}/acts/{api_id}/runs",
            params={"token": APIFY_TOKEN},
            json=payload,
            timeout=30,
        )
        r.raise_for_status()
        run_id = r.json()["data"]["id"]
        print(f"  [{actor_id}] started → {run_id}")
        return run_id
    except Exception as exc:
        print(f"  [{actor_id}] start failed: {exc}")
        return None


async def _wait_run(
    client: httpx.AsyncClient, run_id: str, label: str
) -> bool:
    deadline = time.monotonic() + RUN_TIMEOUT
    while time.monotonic() < deadline:
        await asyncio.sleep(POLL_INTERVAL)
        try:
            r = await client.get(
                f"{APIFY_BASE}/actor-runs/{run_id}",
                params={"token": APIFY_TOKEN},
                timeout=15,
            )
            status = r.json()["data"]["status"]
            if status == "SUCCEEDED":
                return True
            if status in ("FAILED", "ABORTED", "TIMED-OUT"):
                print(f"  [{label}] run ended with status: {status}")
                return False
        except Exception as exc:
            print(f"  [{label}] poll error: {exc}")
    print(f"  [{label}] timed out after {RUN_TIMEOUT}s")
    return False


async def _fetch_items(client: httpx.AsyncClient, run_id: str) -> list:
    try:
        r = await client.get(
            f"{APIFY_BASE}/actor-runs/{run_id}/dataset/items",
            params={"token": APIFY_TOKEN, "format": "json", "limit": 300},
            timeout=60,
        )
        if r.status_code == 200:
            return r.json() or []
    except Exception as exc:
        print(f"  fetch error: {exc}")
    return []


async def run_actor(
    client: httpx.AsyncClient,
    sem: asyncio.Semaphore,
    actor_id: str,
    payload: dict,
    label: str,
) -> list:
    async with sem:
        run_id = await _start_run(client, actor_id, payload)
        if not run_id:
            return []
        ok = await _wait_run(client, run_id, label)
        if not ok:
            return []
        items = await _fetch_items(client, run_id)
        print(f"  [{label}] → {len(items)} items")
        return items


# ---------------------------------------------------------------------------
# Per-source result normalizers
# ---------------------------------------------------------------------------
def _s(d: dict, *keys: str, default: str = "") -> str:
    for k in keys:
        if d.get(k):
            return str(d[k]).strip()
    return default


def _find_url(d: dict) -> str:
    """Fallback: scan all fields and return the first value that starts with http
    and whose key name suggests it is a URL or link.
    Skips image/logo URLs (common false-positive from Naukri/LinkedIn actors)."""
    _IMG_KEY_HINTS = ("logo", "image", "photo", "avatar", "icon", "thumbnail", "banner")
    _IMG_VAL_HINTS = (".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp", ".ico",
                      "img.", "/images/", "/logos/", "naukimg.com")
    for k, v in d.items():
        if v and isinstance(v, str) and v.startswith("http"):
            if any(h in k.lower() for h in ("url", "link", "apply", "href")):
                k_low = k.lower()
                v_low = v.lower()
                if any(s in k_low for s in _IMG_KEY_HINTS):
                    continue
                if any(s in v_low for s in _IMG_VAL_HINTS):
                    continue
                return v.strip()
    return ""


def _title_from_post_text(text: str) -> str:
    for line in text.split("\n"):
        line = line.strip()
        if 10 <= len(line) <= 80 and any(
            w in line.lower()
            for w in ("intern", "engineer", "developer", "hiring", "scientist", "analyst")
        ):
            return line[:80]
    return text[:60]


NORMALIZERS: dict = {
    "linkedin_jobs": lambda x: {
        "title": _s(x, "title", "jobTitle"),
        "company": _s(x, "company", "companyName"),
        "location": _s(x, "location", "jobLocation"),
        "salary": _s(x, "salary", "salaryInfo"),
        "url": _s(x, "applyUrl", "url", "jobUrl", "link") or _find_url(x),
        "description": _s(x, "description", "jobDescription")[:700],
        "source": "LinkedIn Jobs",
        "posted": _s(x, "postedAt", "publishedAt", "date"),
    },
    "linkedin_posts": lambda x: {
        "title": _s(x, "title") or _title_from_post_text(_s(x, "text", "content", "postText")),
        "company": _s(x, "authorName", "authorCompany", "company"),
        "location": _s(x, "location"),
        "salary": "",
        "url": _s(x, "postUrl", "url", "link") or _find_url(x),
        "description": _s(x, "text", "content", "postText")[:700],
        "source": "LinkedIn Post",
        "posted": _s(x, "postedAt", "timestamp", "date"),
    },
    "indeed": lambda x: {
        "title": _s(x, "positionName", "title", "jobTitle"),
        "company": _s(x, "company", "companyName"),
        "location": _s(x, "location", "jobLocation"),
        "salary": _s(x, "salary", "salaryText"),
        "url": _s(x, "url", "jobUrl", "externalApplyLink") or _find_url(x),
        "description": _s(x, "description", "snippet")[:700],
        "source": "Indeed",
        "posted": _s(x, "date", "postedAt"),
    },
    "glassdoor": lambda x: {
        "title": _s(x, "jobTitle", "title"),
        "company": _s(x, "employerName", "company"),
        "location": _s(x, "location", "jobLocation"),
        "salary": _s(x, "payPeriod", "salary", "salaryEstimate"),
        "url": _s(x, "jobViewUrl", "url", "link") or _find_url(x),
        "description": _s(x, "jobDescription", "description")[:700],
        "source": "Glassdoor",
        "posted": _s(x, "listingDate", "postedAt", "date"),
    },
    "naukri": lambda x: {
        "title": _s(x, "jobTitle", "title", "designation"),
        "company": _s(x, "companyName", "company", "ambitionBoxData"),
        "location": _s(x, "location", "jobLocation", "placeholders"),
        "salary": _s(x, "salary", "salaryDetail", "salaryMin"),
        "url": _s(x, "jdURL", "jdUrl", "jobUrl", "jobLink", "applyUrl", "url", "link") or _find_url(x),
        "description": _s(x, "jobDescription", "description", "jobDetails")[:700],
        "source": "Naukri",
        "posted": _s(x, "freshness", "postedAt", "createdDate", "date"),
    },
    "internshala": lambda x: {
        "title": _s(x, "title", "internshipTitle", "profile"),
        "company": _s(x, "company", "companyName", "organizationName"),
        "location": _s(x, "location", "city", "jobLocation") or "India",
        "salary": _s(x, "stipend", "salary", "stipendAmount", "monthlyStipend"),
        "url": _s(x, "url", "link", "applyLink", "apply_link", "internshipUrl") or _find_url(x),
        "description": _s(x, "description", "about", "jobDescription", "details")[:700],
        "source": "Internshala",
        "posted": _s(x, "postedOn", "postedAt", "startDate", "date"),
    },
    "wellfound": lambda x: {
        "title": _s(x, "title", "jobTitle", "role"),
        "company": _s(x, "company", "companyName", "startupName", "organizationName"),
        "location": _s(x, "location", "jobLocation", "remote"),
        "salary": _s(x, "compensation", "salary", "equity", "salaryRange"),
        "url": _s(x, "url", "jobUrl", "link", "applyUrl") or _find_url(x),
        "description": _s(x, "description", "jobDescription", "about")[:700],
        "source": "Wellfound",
        "posted": _s(x, "postedAt", "createdAt", "date"),
    },
}


# ---------------------------------------------------------------------------
# Filters
# ---------------------------------------------------------------------------
def is_india_job(job: dict) -> bool:
    # Internshala and Naukri are India-only platforms — always pass
    if job.get("source") in ("Internshala", "Naukri"):
        return True
    loc = (job.get("location") or "").lower()
    desc = (job.get("description") or "").lower()
    if any(kw in loc for kw in NON_INDIA_LOCS):
        return False
    if any(kw in loc for kw in INDIA_LOCS):
        return True
    if not loc:
        return any(kw in desc for kw in INDIA_LOCS)
    return False


# ---------------------------------------------------------------------------
# Freshness filter — drop jobs posted more than MAX_JOB_AGE_DAYS ago
# ---------------------------------------------------------------------------
MAX_JOB_AGE_DAYS = 45

_DATE_FMTS = ("%Y-%m-%d", "%d %b %Y", "%b %d, %Y", "%d/%m/%Y", "%Y/%m/%d")


def is_recent_job(job: dict, max_days: int = MAX_JOB_AGE_DAYS) -> bool:
    """Return True if the job appears recent, or if we cannot determine its age."""
    raw = (job.get("posted") or "").lower().strip()
    if not raw:
        return True  # no date info — keep
    # Clearly fresh signals
    if any(w in raw for w in ("just", "today", "fresh", "hour", "minute", "second")):
        return True
    # "X days ago"
    m = re.search(r"(\d+)\s*day", raw)
    if m:
        return int(m.group(1)) <= max_days
    # "X weeks ago"
    m = re.search(r"(\d+)\s*week", raw)
    if m:
        return int(m.group(1)) * 7 <= max_days
    # "X months ago" — 1 month is borderline, 2+ months → drop
    m = re.search(r"(\d+)\s*month", raw)
    if m:
        return int(m.group(1)) * 30 <= max_days
    # Any mention of "year" → definitely stale
    if "year" in raw:
        return False
    # Try absolute date formats
    for fmt in _DATE_FMTS:
        try:
            dt = datetime.strptime(raw, fmt)
            return (datetime.now() - dt).days <= max_days
        except ValueError:
            pass
    return True  # unrecognised format — keep to avoid false drops


_LI_POST_JOB_HINTS = (
    "intern", "hiring", "role", "position", "developer", "engineer",
    "apply", "opening", "opportunity", "vacancy", "job",
)


def passes_role_filter(job: dict, profile: dict) -> bool:
    exclude = [
        w.lower()
        for w in profile.get("job_preferences", {}).get("exclude_keywords", [])
    ]
    combined = (job.get("title", "") + " " + job.get("description", "")).lower()
    if any(w in combined for w in exclude):
        return False
    # LinkedIn Posts extra: drop posts that don't mention any job-related keyword
    if job.get("source") == "LinkedIn Post":
        return any(h in combined for h in _LI_POST_JOB_HINTS)
    return True


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
_H_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
_H_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADERS = [
    "#", "Title", "Company", "Location",
    "Salary / Stipend", "Source", "Posted", "Description", "URL",
]
_WIDTHS = [4, 38, 25, 22, 20, 16, 14, 65, 55]


def export_to_excel(jobs: list, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    for col, (h, w) in enumerate(zip(_HEADERS, _WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _H_FONT
        cell.fill = _H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, job in enumerate(jobs, 1):
        row_vals = [
            i,
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary", ""),
            job.get("source", ""),
            job.get("posted", ""),
            job.get("description", ""),
            job.get("url", ""),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val or "")
            cell.alignment = Alignment(wrap_text=(col == 8), vertical="top")
            if i % 2 == 0:
                cell.fill = _ALT_FILL
            if col == 9 and val:  # URL column — make it a clickable hyperlink
                cell.hyperlink = str(val)
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Saved {len(jobs)} jobs → {path}")


# ---------------------------------------------------------------------------
# Claude AI prompt builder
# ---------------------------------------------------------------------------
def build_claude_prompt(profile: dict, resume_text: str, job_count: int) -> str:
    p        = profile.get("personal", {})
    edu      = profile.get("education", {})
    prefs    = profile.get("job_preferences", {})
    skills   = profile.get("skills", {})
    projects = profile.get("projects", [])
    keywords = profile.get("search", {}).get("keywords", [])
    i_cats   = profile.get("search", {}).get(
        "internshala_categories", ["Software Development", "Machine Learning"]
    )

    all_skills = (
        skills.get("primary", [])
        + skills.get("secondary", [])
        + skills.get("ml_ai", [])
    )
    project_lines = "\n".join(
        f"  • {proj['name']}: {proj.get('tech', '')} — {proj.get('description', '')}"
        for proj in projects
    )
    resume_snippet = (
        resume_text[:2500] if resume_text
        else "(resume not loaded — please attach the PDF)"
    )

    roles_str     = ", ".join(prefs.get("roles", []))
    locations_str = ", ".join(prefs.get("preferred_locations", ["India"]))
    skills_str    = ", ".join(all_skills)
    name          = p.get("name", "")
    degree        = edu.get("degree", "")
    institution   = edu.get("institution", "")
    year          = edu.get("year", "")
    gpa           = edu.get("gpa", "")
    location      = p.get("location", "")
    job_type      = prefs.get("type", "internship")

    return f"""You have two files attached:
  1. **jobs.xlsx** — {job_count} internship listings scraped from LinkedIn, Naukri, and Internshala (India, pre-filtered)
  2. **Resume PDF** — my full CV

Complete BOTH parts below before writing ANY output. Do not skip Part 2.

{'='*64}
PART 1 — SCORE EVERY ROW IN THE EXCEL FILE
{'='*64}
Read each row silently. Score it 0–10 using the rubric further below.
Keep all scores in working memory — you will merge them with Part 2 results.

{'='*64}
PART 2 — GOOGLE SEARCH FOR ADDITIONAL LIVE JOBS (posted this week)
{'='*64}
Use your **web search tool** (NOT browser navigation) to run each query below as a Google search.
For each result in the search snippets, extract the job Title, Company, Location, Stipend, and the EXACT URL shown in the search result.

⚠️  STRICT URL RULE FOR PART 2:
- Copy the EXACT URL from the Google search result snippet — do NOT modify, reconstruct, or guess any URL
- If the search snippet does not show a direct apply URL, skip that listing entirely
- NEVER visit a job board and manually construct a URL (e.g. internshala.com/internship/detail/[guessed-id])
- A company careers homepage (e.g. jobs.natwestgroup.com/search/jobs) is NOT a valid apply URL — skip it

### Google searches to run (run all of these):
  1. site:internshala.com/internship/detail python internship paid work-from-home 2025
  2. site:internshala.com/internship/detail machine learning internship stipend India 2025
  3. site:internshala.com/internship/detail fastapi OR backend internship India stipend
  4. site:unstop.com python OR "machine learning" internship India 2025 paid
  5. site:indeed.co.in python developer intern freshers India 2025
  6. site:linkedin.com/jobs/view python backend intern India 2025
  7. site:linkedin.com/jobs/view "machine learning" intern India remote 2025

{'='*64}
MY PROFILE  (use this for ALL scoring — both Excel rows and web-found jobs)
{'='*64}
Name            : {name}
Education       : {degree} at {institution} | Year {year} | GPA {gpa}
Current city    : {location}
Open to         : {locations_str}  ← **STRONG preference for Remote or Hyderabad. Non-remote outside these = lower score.**
Looking for     : {job_type} — PAID only (stipend must be stated or company must be top-tier)
Home base       : Jabalpur, MP — relocating to a metro city costs ₹8K–15K/month just for rent + food
Target roles    : {roles_str}

Skills          : {skills_str}

Projects:
{project_lines}

Resume (first 2500 chars — full PDF attached):
{resume_snippet}

{'='*64}
SCORING RUBRIC  (0–10)
{'='*64}
10  Python / FastAPI / Django backend intern · Remote OR Hyderabad · paid · stipend decent
 9  ML / AI / Data Science intern · Remote OR Hyderabad · paid · stipend stated
 8  Full-stack (React + Python/Node) intern · Remote OR Hyderabad · paid
 7  Same roles but in Pune / Chennai / other tier-1 city with livable stipend (see minimums below)
 6  Bangalore / Mumbai / Delhi role — ONLY if stipend ≥ ₹20,000/month (rent alone is ₹12K+)
 4  Requires 2–3 yrs  OR  unpaid but top-tier company (Google, Microsoft, unicorn)
 2  Stipend unknown / unrecognised company / non-preferred city with borderline pay
 0  Non-India · non-tech · requires 3+ yrs · spam  →  DROP immediately

MINIMUM LIVABLE STIPEND BY CITY (deduct 2 points if below these; drop if far below AND company unknown):
  Remote (WFH)  : ₹5,000+/month  (living at home, no rent)
  Hyderabad     : ₹12,000+/month
  Pune          : ₹15,000+/month
  Chennai       : ₹15,000+/month
  Bengaluru     : ₹20,000+/month  (most expensive — ₹10–15K is NOT livable here)
  Mumbai / Delhi: ₹22,000+/month  (very high cost of living)

{'='*64}
HARD FILTERS — silently drop any job that matches these
{'='*64}
✗  Location outside India (except explicitly "Remote — open to India")
✗  Requires 3 or more years of experience
✗  Unpaid AND company is not widely recognised
✗  Non-tech role (marketing, HR, sales, operations, content writing)
✗  Duplicate URLs (keep only the first occurrence)
✗  Spam / "earn from home" / multi-level / vague "work on exciting projects"
✗  No direct apply URL found anywhere (Excel URL column empty AND not findable via web search) → drop
✗  Bengaluru / Mumbai / Delhi office role with stipend < ₹15,000/month AND unknown company → drop
    (₹10–15K in Bengaluru doesn't cover rent ₹8–12K + food ₹5K + transport ₹2K = ₹15–19K baseline)

{'='*64}
FINAL OUTPUT — ONE UNIFIED RANKED LIST  (Excel jobs + web-found jobs merged)
{'='*64}
Deduplicate across both sources. Sort by score descending within each tier.
Show a source tag [Excel] or [Web] on each card.

---

### 🔥 Tier 1 — Apply Today  (score 8–10)
Best-fit: Python/FastAPI backend · ML/AI · full-stack — India — paid — stipend stated

> **[Job Title]** · [Company] · [City / Remote]
> Stipend: ₹X/month · Score: X/10 · Source: [Excel|Web]
> Why it fits: [one sentence referencing my actual skills / projects]
> 🔗 Apply: <full direct URL here>  ← THIS LINE IS MANDATORY. If no URL exists, skip the job entirely.

### ✅ Tier 2 — Strong Match  (score 6–7)
Good fit: cloud, SWE, adjacent tech — India — paid

(same card format)

### 📋 Tier 3 — Bulk Apply  (score 4–5)
Decent fit, legitimate company, low effort to apply

(same card format)

---

End with exactly this summary line:
`Excel: {job_count} scraped | Web-found: W new | After filters: X total | Tier 1: A | Tier 2: B | Tier 3: C`

**Important rules for the output**:
- **Every job card MUST have a working direct apply URL on the 🔗 Apply line — no exceptions**
- If the Excel URL column is empty and you cannot find the apply link via web search, drop the job
- Never use a company homepage, LinkedIn company page, or search results page as the URL
- If a LinkedIn Post listing has no direct apply URL visible in the description, drop it
- Prefer Remote or Hyderabad roles when scores are tied
- Zero tolerance for fake internship farms or unverified "apply via WhatsApp" listings

**CRITICAL — Web-found URLs must come from Google search result snippets. NO hallucination allowed:**
- For ANY web-found listing, the URL MUST be copied verbatim from the Google search result snippet
- You do NOT need to browse to or click the URL to verify it — the Google search snippet is sufficient
- Job boards like Internshala, Naukri, and LinkedIn show their individual listing URLs in Google results; copy those exactly
- NEVER construct or guess a URL by inserting a job ID or slug you inferred (e.g. internshala.com/internship/detail/XXXXXX)
- NEVER reuse a URL pattern from one listing and substitute different IDs or titles
- If no direct listing URL appears in the search snippet (only a homepage or category page), skip that listing
- It is far better to show 5 real web-found listings than 15 where URLs were guessed
"""


# ---------------------------------------------------------------------------
# Actor orchestration
# ---------------------------------------------------------------------------
async def collect_jobs(
    client: httpx.AsyncClient,
    profile: dict,
    sem: asyncio.Semaphore,
) -> list:
    actor_cfg = profile.get("apify_actors", {})
    keywords = profile.get("search", {}).get(
        "keywords", ["python backend intern India"]
    )

    tasks: list[tuple[str, str, dict]] = []

    def queue(source_key: str, extra_payload: dict) -> None:
        cfg = actor_cfg.get(source_key, {})
        if not cfg.get("enabled", True):
            return
        actor_id = cfg.get("actor_id", "").strip()
        if not actor_id:
            print(f"  Skipping {source_key}: no actor_id configured in user_profile.yaml")
            return
        merged = {**cfg.get("input_template", {}), **extra_payload}
        tasks.append((source_key, actor_id, merged))

    # LinkedIn Jobs — very cheap (~$0.01/25 results); use all keywords for max coverage
    for kw in keywords:
        li_url = (
            "https://www.linkedin.com/jobs/search/"
            f"?keywords={quote_plus(kw)}&location=India&f_TPR=r604800&position=1&pageNum=0"  # f_TPR=r604800 = past 7 days
        )
        queue("linkedin_jobs", {"urls": [li_url]})

    # LinkedIn Posts — free actor; use all keywords for max coverage
    for kw in keywords:
        queue("linkedin_posts", {"query": f"hiring {kw} India"})

    # Indeed India — most expensive actor ($6/1K); 1 run only to avoid 402 on free tier
    for kw in keywords[:1]:
        queue("indeed", {"position": kw})

    # Naukri — field is "keyword" (not "position", confirmed from actor JSON schema)
    for kw in keywords[:2]:
        queue("naukri", {"keyword": kw})

    # Internshala — category-based (not keyword), uses internshala_categories from profile
    internshala_cats = (
        profile.get("search", {})
        .get("internshala_categories", ["Software Development"])
    )
    for cat in internshala_cats[:3]:
        queue("internshala", {"job_category": cat})

    # Wellfound — crawlerbros actor takes keyword + location
    for kw in keywords[:2]:
        queue("wellfound", {"keyword": kw})

    # Glassdoor
    for kw in keywords[:2]:
        queue("glassdoor", {"keyword": kw})

    print(f"\nQueued {len(tasks)} actor runs...\n")

    coros = [
        run_actor(client, sem, actor_id, payload, f"{src}:{idx}")
        for idx, (src, actor_id, payload) in enumerate(tasks)
    ]
    results = await asyncio.gather(*coros, return_exceptions=True)

    all_jobs: list[dict] = []
    for (source_key, _, _), result in zip(tasks, results):
        if isinstance(result, Exception):
            print(f"  Error in {source_key}: {result}")
            continue
        norm = NORMALIZERS.get(source_key, lambda x: x)
        for item in result:
            try:
                all_jobs.append(norm(item))
            except Exception:
                pass

    return all_jobs


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description="Job Finder — Apify-powered India internship scraper"
    )
    ap.add_argument("--profile", default="user_profile.yaml", help="Path to user_profile.yaml")
    ap.add_argument("--resume", default=os.getenv("RESUME_PATH", ""), help="Path to resume PDF")
    ap.add_argument("--output", default=OUTPUT_EXCEL, help="Output Excel filename")
    ap.add_argument(
        "--prompt-only",
        action="store_true",
        help="Skip Apify scraping; regenerate the Claude prompt from the existing Excel file",
    )
    return ap.parse_args()


async def main() -> None:
    args = parse_args()

    if not Path(args.profile).exists():
        print(f"ERROR: Profile file not found: {args.profile}")
        print("  Fill in user_profile.yaml before running.")
        return

    profile = load_profile(args.profile)
    print(f"Profile  : {profile.get('personal', {}).get('name', 'Unknown')}")

    resume_path = args.resume or profile.get("resume_path", "")
    resume_text = read_resume(resume_path) if resume_path else ""
    if resume_text:
        print(f"Resume   : {len(resume_text)} chars from {resume_path}")
    else:
        print("Resume   : not loaded (set resume_path in user_profile.yaml or use --resume)")

    output_path = args.output

    # --prompt-only: skip scraping, just regenerate the Claude prompt from existing Excel
    if args.prompt_only:
        if not Path(output_path).exists():
            print(f"ERROR: --prompt-only requires an existing Excel file at: {output_path}")
            print("  Run without --prompt-only first to scrape and create the file.")
            return
        import openpyxl as _ox
        wb = _ox.load_workbook(output_path, read_only=True)
        job_count = wb.active.max_row - 1  # subtract header row
        wb.close()
        print(f"\nPrompt-only mode: read {job_count} rows from {output_path}")
        prompt = build_claude_prompt(profile, resume_text, job_count)
        prompt_file = Path(output_path).stem + "_claude_prompt.txt"
        Path(prompt_file).write_text(prompt, encoding="utf-8")
        print(f"Claude prompt   → {prompt_file}")
        print("\n" + "=" * 60)
        print("NEXT STEPS — Score your jobs with Claude AI:")
        print("  1. Open https://claude.ai")
        print(f"  2. Attach:  {output_path}")
        if resume_path and Path(resume_path).exists():
            print(f"  3. Attach:  {resume_path}  (your resume PDF)")
        print(f"  4. Paste the contents of:  {prompt_file}")
        print("=" * 60)
        return

    if not APIFY_TOKEN:
        print("ERROR: APIFY_TOKEN is not set.")
        print("  1. Get your token: https://console.apify.com/settings/integrations")
        print("  2. Create a .env file here and add:  APIFY_TOKEN=your_token_here")
        return

    sem = asyncio.Semaphore(MAX_PARALLEL)
    async with httpx.AsyncClient() as client:
        all_jobs = await collect_jobs(client, profile, sem)

    print(f"\nCollected       : {len(all_jobs)}")

    india_jobs = [j for j in all_jobs if is_india_job(j)]
    print(f"India-only      : {len(india_jobs)}")

    filtered = [j for j in india_jobs if passes_role_filter(j, profile)]
    print(f"After role filter: {len(filtered)}")

    filtered = [j for j in filtered if is_recent_job(j)]
    print(f"After date filter : {len(filtered)} (dropped jobs older than {MAX_JOB_AGE_DAYS} days)")

    seen: set[str] = set()
    unique: list[dict] = []
    for j in filtered:
        url = j.get("url", "")
        if url and url in seen:
            continue
        if url:
            seen.add(url)
        unique.append(j)
    print(f"After dedup     : {len(unique)}")

    if not unique:
        print("\nNo jobs found. Check your actor IDs in user_profile.yaml and APIFY_TOKEN.")
        return

    export_to_excel(unique, output_path)

    prompt = build_claude_prompt(profile, resume_text, len(unique))
    prompt_file = Path(output_path).stem + "_claude_prompt.txt"
    Path(prompt_file).write_text(prompt, encoding="utf-8")
    print(f"Claude prompt   → {prompt_file}")

    print("\n" + "=" * 60)
    print("NEXT STEPS — Score your jobs with Claude AI:")
    print("  1. Open https://claude.ai")
    print(f"  2. Attach:  {output_path}")
    if resume_path and Path(resume_path).exists():
        print(f"  3. Attach:  {resume_path}  (your resume PDF)")
    print(f"  4. Paste the contents of:  {prompt_file}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
